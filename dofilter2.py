import random
import time

import openpyxl
import requests
from bs4 import BeautifulSoup

def extract_cve_numbers(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook['漏洞列表']
    # Check if the header contains '漏洞编号'
    header_row = sheet[1]
    header_values = [cell.value for cell in header_row]
    if '漏洞编号' not in header_values:
        print("没有列为 '漏洞编号'.")
        return []
    cve_numbers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cve_number = row[header_values.index('漏洞编号')]
        if cve_number:
            if "CVE" in cve_number:
                cve_numbers.append(cve_number[4:])
                #cve_numbers.append(cve_number)
            else: continue
    workbook.close()
    return cve_numbers

def getPocInformation(CVE_List):
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/113.0.0.0 Safari/537.36",
        "Host": "avd.aliyun.com"
    }
    CVE_res={}
    for CVE_Num in CVE_List:
        time.sleep(random.randint(4,8))
        url = f"https://avd.aliyun.com/detail?id=AVD-{CVE_Num}"
        #print(url)
        res=requests.get(headers=header,url=url)
        soup = BeautifulSoup(res.text, 'html.parser')
        metric_values = soup.find_all("div", class_="metric-value")
        if len(metric_values) >= 2:
            second_value = metric_values[1].text.strip()
        else:
            continue
        if "暂无" in second_value:
            #print(second_value)
            continue
        else:
            CVE_Num = f"CVE-{CVE_Num}"
            CVE_name_value = soup.find("span", class_="header__title__text")
            cve_name = CVE_name_value.text.strip()
            CVE_res[CVE_Num] = cve_name
            print(CVE_res)
    return CVE_res

# def Write_into_file(CVE_res):
#     print(CVE_res)
#     wb = openpyxl.load_workbook("cve_res.xlsx")
#     sheet = wb.active
#     for i, cve in enumerate(CVE_res, start=1):
#         sheet.cell(row=i, column=1, value=cve)
#     wb.save("cve_res.xlsx")
def Write_into_file(CVE_res):
    print(CVE_res)
    wb = openpyxl.load_workbook("cve_res.xlsx")
    sheet = wb.active
    for i, (key, value) in enumerate(CVE_res.items(), start=1):
        sheet.cell(row=i, column=1, value=key)
        sheet.cell(row=i, column=2, value=value)
    wb.save("cve_res.xlsx")

if __name__ == '__main__':
    CVE_List = extract_cve_numbers("漏洞信息报表.xlsx");
    #print(set(CVE_List))
    print(f"共计{len(set(CVE_List))}个编号")
    CVE_res = getPocInformation(set(CVE_List))
    print(f"过滤后共计{len(CVE_res)}个编号")
    Write_into_file(CVE_res)



